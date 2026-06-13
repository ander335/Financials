import csv
import math
import re
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import column_index_from_string
from PIL import Image as PILImage


PRICE_BAND_CELLS = ("K2", "L2", "M2", "N2", "O2")
CURRENCY_SYMBOLS = {
    "USD": "$",
    "EUR": "\u20ac",
    "GBP": "\u00a3",
    "JPY": "\u00a5",
    "CHF": "CHF",
}
CURRENCY_LOCALE_IDS = {
    "USD": "409",
    "EUR": "2",
    "GBP": "809",
    "JPY": "411",
}
A1_REFERENCE_RE = re.compile(
    r"(?:(?:'(?P<quoted_sheet>(?:[^']|'')+)'|(?P<sheet>[A-Za-z_][A-Za-z0-9_.]*))!)?"
    r"(?P<start_col>\$?[A-Z]{1,3})(?P<start_row>\$?\d+)"
    r"(?::(?P<end_col>\$?[A-Z]{1,3})(?P<end_row>\$?\d+))?"
)


def open_workbook(path, keep_vba=True):
    return load_workbook(path, keep_vba=keep_vba)


def save_workbook(wb, output_path, validate=True, keep_vba=True):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    set_recalculate_on_open(wb)
    wb.save(output_path)
    if validate:
        load_workbook(output_path, keep_vba=keep_vba, read_only=True).close()
    return output_path


def image_anchor_row(image):
    anchor = getattr(image, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return marker.row + 1


def remove_images_from_row(ws, min_row):
    remaining = []
    for image in getattr(ws, "_images", []):
        row = image_anchor_row(image)
        if row is None or row < min_row:
            remaining.append(image)
    ws._images = remaining


def add_scaled_png(ws, image_path, anchor_cell="E21", max_width=1214, max_height=221, replace_from_row=18):
    image_path = Path(image_path)
    if not image_path.exists():
        raise FileNotFoundError(image_path)
    if replace_from_row is not None:
        remove_images_from_row(ws, replace_from_row)

    with PILImage.open(image_path) as source:
        width, height = source.size
    if width <= 0 or height <= 0:
        raise ValueError(f"Logo image has invalid dimensions: {image_path}")

    scale = min(max_width / width, max_height / height, 1)
    logo = XLImage(image_path)
    logo.width = int(width * scale)
    logo.height = int(height * scale)
    ws.add_image(logo, anchor_cell)
    return logo.width, logo.height


def read_csv_rows(path):
    with Path(path).open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def detect_single_file(folder, pattern):
    matches = sorted(Path(folder).glob(pattern))
    if len(matches) != 1:
        raise ValueError(f"Expected exactly one {pattern} in {folder}, found {len(matches)}.")
    return matches[0]


def parse_number(value):
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if text == "":
        return None
    number = float(text)
    return int(number) if number.is_integer() else number


def parse_year(value):
    return int(str(value).strip().split()[0])


def parse_iso_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def latest_row_by(rows, column):
    if not rows:
        return None
    return max(rows, key=lambda row: row[column])


def latest_number(rows, sort_column, value_column):
    latest = latest_row_by(rows, sort_column)
    return None if latest is None else parse_number(latest[value_column])


def nice_price_step(current_price):
    target = max(abs(float(current_price)) * 0.2, 1)
    magnitude = 10 ** math.floor(math.log10(target))
    for multiplier in (1, 2, 5, 10):
        step = multiplier * magnitude
        if target <= step:
            return int(step) if float(step).is_integer() else step
    return 10 * magnitude


def clean_price(value):
    rounded = round(value)
    return int(rounded) if math.isclose(value, rounded, rel_tol=0, abs_tol=1e-9) else value


def price_sensitivity_values(current_price, step=None):
    current = float(current_price)
    step = float(step if step is not None else nice_price_step(current))
    display_current = round(current)

    lower_one = math.floor(current / step) * step
    if lower_one >= display_current:
        lower_one -= step
    lower_two = lower_one - step

    upper_one = math.ceil(current / step) * step
    if upper_one <= display_current or upper_one - display_current < step / 2:
        upper_one += step / 2
    upper_two = upper_one + step

    return [
        clean_price(lower_two),
        clean_price(lower_one),
        current_price,
        clean_price(upper_one),
        clean_price(upper_two),
    ]


def populate_result_prices(ws, current_price, current_cell="C2", price_cells=PRICE_BAND_CELLS, step=None):
    ws[current_cell] = current_price
    values = price_sensitivity_values(current_price, step=step)
    if len(price_cells) != len(values):
        raise ValueError(f"Expected {len(values)} price cells, got {len(price_cells)}.")
    for cell, value in zip(price_cells, values):
        ws[cell] = value
    return values


def currency_symbol(currency):
    code = str(currency).strip().upper()
    if not code:
        raise ValueError("Currency is required.")
    return CURRENCY_SYMBOLS.get(code, code)


def currency_locale_id(currency):
    return CURRENCY_LOCALE_IDS.get(str(currency).strip().upper())


def currency_markers(currency):
    code = str(currency).strip().upper()
    markers = {code, currency_symbol(code)}
    locale_id = currency_locale_id(code)
    if locale_id:
        markers.add(f"[${currency_symbol(code)}-{locale_id}]")
    return {marker for marker in markers if marker}


def is_currency_number_format(number_format, source_currency="USD"):
    if not number_format or number_format == "General":
        return False
    code = str(source_currency).strip().upper()
    symbol = currency_symbol(code)
    return (
        code in number_format
        or f'"{symbol}"' in number_format
        or re.search(rf"\[\${re.escape(symbol)}(-[0-9A-Fa-f]+)?\]", number_format) is not None
    )


def convert_currency_number_format(number_format, target_currency, source_currency="USD"):
    if not is_currency_number_format(number_format, source_currency=source_currency):
        return number_format

    source_code = str(source_currency).strip().upper()
    target_code = str(target_currency).strip().upper()
    source_symbol = currency_symbol(source_code)
    target_symbol = currency_symbol(target_code)

    protected_tokens = []

    def protect_token(token):
        placeholder = f"__CURRENCY_FORMAT_{len(protected_tokens)}__"
        protected_tokens.append((placeholder, token))
        return placeholder

    updated = re.sub(
        rf"\[\${re.escape(source_symbol)}(-[0-9A-Fa-f]+)?\]",
        lambda _match: f'"{target_symbol}"',
        number_format,
    )
    updated = re.sub(
        r"\[\$[^]]+\]",
        lambda match: protect_token(match.group(0)),
        updated,
    )
    updated = updated.replace(source_code, target_code)
    updated = updated.replace(f'"{source_symbol}"', f'"{target_symbol}"')
    for placeholder, replacement in protected_tokens:
        updated = updated.replace(placeholder, replacement)
    return updated


def apply_currency_number_formats(wb, target_currency, source_currency="USD", sheets=None):
    sheet_names = sheets or wb.sheetnames
    changes = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                old_format = cell.number_format
                new_format = convert_currency_number_format(
                    old_format,
                    target_currency=target_currency,
                    source_currency=source_currency,
                )
                if new_format != old_format:
                    cell.number_format = new_format
                    changes.append(
                        {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "old_format": old_format,
                            "new_format": new_format,
                        }
                    )
    return changes


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.protection:
        target.protection = copy(source.protection)


def copy_row_styles(ws, source_row, target_row, first_col, last_col):
    for col in range(first_col, last_col + 1):
        copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def insert_styled_rows(ws, row, amount=1, source_row=None, first_col=1, last_col=None):
    if amount < 1:
        raise ValueError("amount must be at least 1")
    last_col = last_col or ws.max_column
    ws.insert_rows(row, amount)

    if source_row is not None:
        source_row = source_row + amount if source_row >= row else source_row
        for target_row in range(row, row + amount):
            copy_row_styles(ws, source_row, target_row, first_col, last_col)
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def translate_row_formulas(ws, source_row, target_row, first_col=1, last_col=None):
    last_col = last_col or ws.max_column
    for column in range(first_col, last_col + 1):
        source = ws.cell(source_row, column)
        if isinstance(source.value, str) and source.value.startswith("="):
            target = ws.cell(target_row, column)
            target.value = Translator(
                source.value,
                origin=source.coordinate,
            ).translate_formula(target.coordinate)


def formula_row_references(formula):
    references = []
    for match in A1_REFERENCE_RE.finditer(formula):
        start_row = int(match.group("start_row").replace("$", ""))
        end_row_text = match.group("end_row")
        end_row = int(end_row_text.replace("$", "")) if end_row_text else start_row
        quoted_sheet = match.group("quoted_sheet")
        references.append(
            {
                "sheet": quoted_sheet.replace("''", "'") if quoted_sheet else match.group("sheet"),
                "min_row": min(start_row, end_row),
                "max_row": max(start_row, end_row),
                "reference": match.group(0),
            }
        )
    return references


def find_formulas_referencing_rows(
    wb,
    min_row,
    max_row,
    target_sheet=None,
    search_sheets=None,
):
    matches = []
    sheet_names = search_sheets or wb.sheetnames
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                for reference in formula_row_references(formula):
                    reference_sheet = reference["sheet"] or sheet_name
                    if target_sheet and reference_sheet != target_sheet:
                        continue
                    if reference["max_row"] < min_row or reference["min_row"] > max_row:
                        continue
                    matches.append(
                        {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": formula,
                            "reference": reference,
                        }
                    )
    return matches


def replace_formula_references(wb, replacements, sheets=None):
    changes = []
    for sheet_name in sheets or wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                formula = cell.value
                if not isinstance(formula, str) or not formula.startswith("="):
                    continue
                updated = formula
                for old_reference, new_reference in replacements.items():
                    updated = updated.replace(old_reference, new_reference)
                if updated != formula:
                    cell.value = updated
                    changes.append(
                        {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "old_formula": formula,
                            "new_formula": updated,
                        }
                    )
    return changes


def compare_metric_records(expected, actual, tolerances=None, fields=None):
    tolerances = tolerances or {}
    fields = fields or sorted(set(expected) | set(actual))
    mismatches = []

    for field in fields:
        expected_value = expected.get(field)
        actual_value = actual.get(field)
        tolerance = tolerances.get(field, 0)

        if expected_value is None or actual_value is None:
            matches = expected_value == actual_value
        elif isinstance(expected_value, (int, float)) and isinstance(actual_value, (int, float)):
            matches = math.isclose(
                expected_value,
                actual_value,
                rel_tol=0,
                abs_tol=tolerance,
            )
        else:
            matches = expected_value == actual_value

        if not matches:
            mismatches.append(
                {
                    "field": field,
                    "expected": expected_value,
                    "actual": actual_value,
                    "tolerance": tolerance,
                }
            )

    return mismatches


def scan_broken_formula_references(wb, sheets=None):
    matches = []
    for sheet_name in sheets or wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                formula = cell.value
                if isinstance(formula, str) and formula.startswith("=") and "#REF!" in formula.upper():
                    matches.append(
                        {
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": formula,
                        }
                    )
    return matches


def scan_circular_formula_references(wb, sheets=None, max_range_cells=10000):
    sheet_names = sheets or wb.sheetnames
    formula_cells = {}
    formula_positions = {}
    dependencies = {}

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                formula = cell.value
                if isinstance(formula, str) and formula.startswith("="):
                    node = (sheet_name, cell.coordinate)
                    formula_cells[node] = formula
                    formula_positions[node] = (cell.row, cell.column)

    for node, formula in formula_cells.items():
        sheet_name, _ = node
        node_dependencies = set()
        for match in A1_REFERENCE_RE.finditer(formula):
            quoted_sheet = match.group("quoted_sheet")
            reference_sheet = (
                quoted_sheet.replace("''", "'")
                if quoted_sheet
                else match.group("sheet") or sheet_name
            )
            if reference_sheet not in sheet_names:
                continue

            start_col = column_index_from_string(
                match.group("start_col").replace("$", "")
            )
            end_col_text = match.group("end_col")
            end_col = (
                column_index_from_string(end_col_text.replace("$", ""))
                if end_col_text
                else start_col
            )
            start_row = int(match.group("start_row").replace("$", ""))
            end_row_text = match.group("end_row")
            end_row = (
                int(end_row_text.replace("$", ""))
                if end_row_text
                else start_row
            )

            min_col, max_col = sorted((start_col, end_col))
            min_row, max_row = sorted((start_row, end_row))
            range_size = (max_col - min_col + 1) * (max_row - min_row + 1)
            if range_size > max_range_cells:
                for dependency, (row, column) in formula_positions.items():
                    if dependency[0] != reference_sheet:
                        continue
                    if min_row <= row <= max_row and min_col <= column <= max_col:
                        node_dependencies.add(dependency)
                continue

            ws = wb[reference_sheet]
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    dependency = (reference_sheet, ws.cell(row, column).coordinate)
                    if dependency in formula_cells:
                        node_dependencies.add(dependency)
        dependencies[node] = node_dependencies

    index = 0
    stack = []
    on_stack = set()
    indices = {}
    low_links = {}
    cycles = []

    def visit(node):
        nonlocal index
        indices[node] = index
        low_links[node] = index
        index += 1
        stack.append(node)
        on_stack.add(node)

        for dependency in dependencies.get(node, ()):
            if dependency not in indices:
                visit(dependency)
                low_links[node] = min(low_links[node], low_links[dependency])
            elif dependency in on_stack:
                low_links[node] = min(low_links[node], indices[dependency])

        if low_links[node] != indices[node]:
            return

        component = []
        while True:
            member = stack.pop()
            on_stack.remove(member)
            component.append(member)
            if member == node:
                break

        is_self_reference = (
            len(component) == 1
            and component[0] in dependencies.get(component[0], ())
        )
        if len(component) > 1 or is_self_reference:
            cycles.append(component)

    for node in formula_cells:
        if node not in indices:
            visit(node)

    results = []
    for component in cycles:
        members = sorted(component)
        results.append(
            {
                "cells": [
                    {
                        "sheet": sheet_name,
                        "cell": cell_ref,
                        "formula": formula_cells[(sheet_name, cell_ref)],
                    }
                    for sheet_name, cell_ref in members
                ]
            }
        )
    return sorted(
        results,
        key=lambda result: (
            result["cells"][0]["sheet"],
            result["cells"][0]["cell"],
        ),
    )


def list_external_links(wb):
    links = []
    for index, link in enumerate(getattr(wb, "_external_links", []), start=1):
        links.append(
            {
                "index": index,
                "file_link": getattr(link, "file_link", None),
            }
        )
    return links


def missing_years(years):
    normalized = sorted(set(int(year) for year in years))
    if len(normalized) < 2:
        return []
    return [
        year
        for year in range(normalized[0], normalized[-1] + 1)
        if year not in normalized
    ]


def classify_period_rows(ws, min_row, max_row, classifiers):
    matches = []
    for row in range(min_row, max_row + 1):
        for category, classifier in classifiers.items():
            if classifier(ws, row):
                matches.append({"row": row, "category": category})
                break
    return matches


def set_recalculate_on_open(wb):
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


def clear_columns(ws, columns, start_row, end_row):
    for row in range(start_row, end_row + 1):
        for column in columns:
            ws[f"{column}{row}"] = None


def write_mapped_rows(ws, rows, start_row, mapping, limit=None):
    written = 0
    source_rows = rows[:limit] if limit else rows
    for row_offset, csv_row in enumerate(source_rows):
        excel_row = start_row + row_offset
        for target_col, spec in mapping.items():
            ws[f"{target_col}{excel_row}"] = resolve_mapping_value(spec, csv_row, excel_row)
        written += 1
    return written


def resolve_mapping_value(spec, csv_row, excel_row):
    if callable(spec):
        return spec(csv_row, excel_row)
    if isinstance(spec, tuple):
        csv_column, converter = spec
        return converter(csv_value(csv_row, csv_column))
    return csv_value(csv_row, spec)


def csv_value(csv_row, column):
    if column in csv_row:
        return csv_row[column]
    normalized_column = normalize_header(column)
    for header, value in csv_row.items():
        if normalize_header(header) == normalized_column:
            return value
    raise KeyError(column)


def normalize_header(header):
    text = str(header).replace("&", "_and_")
    normalized = "".join(
        char.lower() if char.isalnum() else "_"
        for char in text
    )
    return "_".join(part for part in normalized.split("_") if part)


def write_columnar_rows(ws, rows, start_row, columns, limit=None):
    mapping = {target_col: spec for target_col, spec in columns}
    return write_mapped_rows(ws, rows, start_row, mapping, limit=limit)


def write_formulas(ws, formulas_by_col, start_row, row_count):
    for row in range(start_row, start_row + row_count):
        for column, formula_factory in formulas_by_col.items():
            value = formula_factory(row) if callable(formula_factory) else formula_factory
            ws[f"{column}{row}"] = value


def set_workbook_cells(wb, updates):
    changes = []
    for sheet_name, cell_updates in updates.items():
        ws = wb[sheet_name]
        for cell_ref, new_value in cell_updates.items():
            cell = ws[cell_ref]
            old_value = cell.value
            cell.value = new_value
            changes.append(
                {
                    "sheet": sheet_name,
                    "cell": cell_ref,
                    "old_value": old_value,
                    "new_value": new_value,
                }
            )
    return changes


def write_two_column_series(ws, rows, start_row, first_col, second_col, first_spec, second_spec, limit=None):
    source_rows = rows[:limit] if limit else rows
    for row_offset, row in enumerate(source_rows):
        excel_row = start_row + row_offset
        ws[f"{first_col}{excel_row}"] = resolve_mapping_value(first_spec, row, excel_row)
        ws[f"{second_col}{excel_row}"] = resolve_mapping_value(second_spec, row, excel_row)
    return len(source_rows)


def assert_same_number_format(ws, source_ref, target_refs):
    expected = ws[source_ref].number_format
    mismatches = []
    for ref in target_refs:
        actual = ws[ref].number_format
        if actual != expected:
            mismatches.append((ref, expected, actual))
    if mismatches:
        details = "; ".join(f"{ref}: expected {expected!r}, got {actual!r}" for ref, expected, actual in mismatches)
        raise AssertionError(details)


def compare_cell_styles(ws, source_ref, target_refs):
    expected = ws[source_ref].style_id
    mismatches = []
    for ref in target_refs:
        actual = ws[ref].style_id
        if actual != expected:
            mismatches.append(
                {
                    "cell": ref,
                    "expected_style_id": expected,
                    "actual_style_id": actual,
                }
            )
    return mismatches
